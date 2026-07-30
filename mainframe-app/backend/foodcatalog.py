"""Open food data → the Mainframe's permanent food catalogue.

Why this exists
---------------
`NUTRITION_SPEC` says your personal library (`:Food`) starts EMPTY and is never
seeded — that rule stands. But an empty library means every food has to be typed
in by hand, which is the one thing MyFitnessPal does that this app didn't.

So there are **two** tables, deliberately:

  * **`:CatalogFood`** — a reference catalogue of real foods with real numbers,
    imported from the sources below. Read-only. Millions of rows. Never appears
    in "my foods".
  * **`:Food`** — your personal library. Still starts empty, still sorts
    most-used-first, still only holds things you actually eat. Adopting a
    catalogue entry *copies* a snapshot into it.

Keeping them apart is the whole design: millions of reference rows would drown a
personal library, and the spec's rule would die quietly. Adoption is the bridge.

Sources
-------
  * **USDA FoodData Central** (public domain) — SR Legacy, Foundation, FNDDS
    survey foods, and the ~1.9M-row Branded set.
  * **Open Food Facts** (Open Database Licence) — crowd-sourced packaged
    products worldwide, with barcodes. The only free source that knows what is
    on a British supermarket shelf.
  * **UK CoFID** (Open Government Licence) — McCance & Widdowson's tables. Small
    but authoritative, and the only one that knows what a digestive biscuit is.

Whole foods are ranked above packaged ones at query time (see routes/nutrition),
because otherwise a search for "banana" returns several hundred cereal bars.

Local-first
-----------
Every source is openly licensed and stored under
`~/.mainframe/foodcatalog/source/`. Once the files are there the import needs no
network at all, ever — re-running it offline works. Attribution for the two
licensed sources: Open Food Facts is ODbL, CoFID is © Crown copyright under the
Open Government Licence.

Data notes that cost time to discover
-------------------------------------
  * **Survey/FNDDS states nutrients by `nutrient_nbr`, not `nutrient_id`.** Its
    `food_nutrient.nutrient_id` column holds 208, 203, 301… where SR Legacy and
    Foundation hold 1008, 1003, 1087. Map only one way and 5,432 foods import
    with every nutrient blank. The lookup below is keyed on both.
  * **The Foundation archive is mostly not Foundation food.** Its `food.csv`
    carries ~88,000 rows of `sample_food` / `sub_sample_food` /
    `market_acquisition` laboratory records around ~469 real foods. Filter on
    `data_type` or you import the lab bookkeeping.
  * **Energy appears more than once** — kcal (1008), kJ (1062) and the two
    Atwater estimates (2047/2048). Only KCAL is wanted, with the Atwater figures
    as a fallback for foods that lack 1008.

Everything USDA states is **per 100 g**. That is stored as-is; the per-serving
conversion the spec demands happens at adoption, where a portion is chosen.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import zipfile
from pathlib import Path

log = logging.getLogger("synapse.foodcatalog")

# --------------------------------------------------------------------------- #
# Where the archives live
# --------------------------------------------------------------------------- #
def source_dir() -> Path:
    from config import settings
    return Path(settings.foodcatalog_source).expanduser()


# Which archive supplies what. `data_type` is the filter applied to food.csv.
SOURCES = [
    ("sr_legacy", "USDA SR Legacy", "sr_legacy_food", "FoodData_Central_sr_legacy_food_csv"),
    ("foundation", "USDA Foundation Foods", "foundation_food", "FoodData_Central_foundation_food_csv"),
    ("survey", "USDA FNDDS (survey)", "survey_fndds_food", "FoodData_Central_survey_food_csv"),
    ("branded", "USDA Branded", "branded_food", "FoodData_Central_branded_food_csv"),
]

OFF_LABEL = "Open Food Facts"

# --------------------------------------------------------------------------- #
# Nutrient mapping
# --------------------------------------------------------------------------- #
# canonical key -> (modern nutrient ids, legacy nutrient_nbrs, required unit or None)
# Order within `ids` is preference order: the first one a food has, wins.
NUTRIENTS: dict[str, tuple[list[str], list[str], str | None]] = {
    "calories":      (["1008", "2047", "2048"], ["208", "957", "958"], "KCAL"),
    "protein":       (["1003"], ["203"], None),
    "fat":           (["1004"], ["204"], None),
    "carbs":         (["1005"], ["205"], None),
    "fibre":         (["1079"], ["291"], None),
    "sugar":         (["2000", "1063"], ["269"], None),
    "saturated_fat": (["1258"], ["606"], None),
    "cholesterol":   (["1253"], ["601"], None),
    "sodium":        (["1093"], ["307"], None),
    "potassium":     (["1092"], ["306"], None),
    "calcium":       (["1087"], ["301"], None),
    "iron":          (["1089"], ["303"], None),
    "magnesium":     (["1090"], ["304"], None),
    "zinc":          (["1095"], ["309"], None),
    "vitamin_c":     (["1162"], ["401"], None),
    "vitamin_a":     (["1106"], ["320"], None),
    "vitamin_d":     (["1114"], ["328"], None),
    "vitamin_b12":   (["1178"], ["418"], None),
    "folate":        (["1177"], ["417"], None),
}

NUTRIENT_KEYS = list(NUTRIENTS)
MACRO_KEYS = ["calories", "protein", "carbs", "fat"]


def _num(v, default=None):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _fmt(n: float) -> str:
    """1.0 -> '1', 0.5 -> '0.5' — portion labels should read like a recipe."""
    return str(int(n)) if float(n).is_integer() else f"{n:g}"


def _open_csv(zf: zipfile.ZipFile, names: dict[str, str], base: str):
    """Yield dict rows from a member CSV, or nothing if the archive lacks it."""
    if base not in names:
        return
    with zf.open(names[base]) as fh:
        yield from csv.DictReader(io.TextIOWrapper(fh, "utf-8-sig"))


def _nutrient_lookup(zf, names) -> dict[str, tuple[str, int]]:
    """Map every id AND legacy nbr a dataset might use -> (canonical key, preference).

    Modern ids win over legacy numbers when both would resolve, which is why the
    id pass runs first and the nbr pass refuses to overwrite.
    """
    rows = list(_open_csv(zf, names, "nutrient.csv"))
    by_id = {r["id"]: r for r in rows}
    lookup: dict[str, tuple[str, int]] = {}

    for key, (ids, nbrs, unit) in NUTRIENTS.items():
        for pref, nid in enumerate(ids):
            row = by_id.get(nid)
            if not row:
                continue
            if unit and (row.get("unit_name") or "").upper() != unit:
                continue          # 1062 is Energy in kJ — not what a calorie is
            lookup[nid] = (key, pref)
        for pref, nbr in enumerate(nbrs):
            if nbr in lookup:     # never let a legacy number shadow a real id
                continue
            # Confirm the number belongs to the nutrient we think it does, and
            # carries the right unit, before trusting it.
            match = next((r for r in rows if r.get("nutrient_nbr") == nbr), None)
            if not match:
                continue
            if unit and (match.get("unit_name") or "").upper() != unit:
                continue
            lookup[nbr] = (key, pref)
    return lookup


def _portion_label(amount, unit_name: str, description: str, modifier: str) -> str:
    """Turn FDC's four half-filled portion columns into something a human reads.

    The three datasets each fill these differently:
      SR Legacy   amount=1, unit=undetermined, modifier='serving'  -> "1 serving"
      Foundation  amount=2, unit='tbsp'                            -> "2 tbsp"
      FNDDS       amount='', description='1 cup', modifier='10205' -> "1 cup"
    A numeric modifier is FNDDS's internal code, never a label.
    """
    desc = (description or "").strip()
    mod = (modifier or "").strip()
    amt = _num(amount)
    unit = (unit_name or "").strip()
    if unit.lower() in ("undetermined", ""):
        unit = ""

    if desc:
        # FNDDS descriptions usually already lead with their own count.
        if amt and not desc[0].isdigit():
            return f"{_fmt(amt)} {desc}"
        return desc

    if not unit and mod and not mod.replace(".", "").isdigit():
        unit = mod
        mod = ""
    if not unit:
        unit = "serving"

    label = f"{_fmt(amt) if amt else '1'} {unit}".strip()
    if mod and not mod.replace(".", "").isdigit():
        label = f"{label}, {mod}"
    return label


def parse_archive(path: Path, data_type: str, source_key: str) -> list[dict]:
    """Read one FDC archive into catalogue records. Pure parsing — no database."""
    with zipfile.ZipFile(path) as zf:
        names = {os.path.basename(n): n for n in zf.namelist() if not n.endswith("/")}

        categories = {r["id"]: r["description"]
                      for r in _open_csv(zf, names, "food_category.csv")}
        # FNDDS foods point at a WWEIA category, not a food_category — without
        # this every survey food imports with a blank category.
        categories.update({r["wweia_food_category"]: r["wweia_food_category_description"]
                           for r in _open_csv(zf, names, "wweia_food_category.csv")})
        units = {r["id"]: r["name"] for r in _open_csv(zf, names, "measure_unit.csv")}
        lookup = _nutrient_lookup(zf, names)

        foods: dict[str, dict] = {}
        for r in _open_csv(zf, names, "food.csv"):
            if r.get("data_type") != data_type:
                continue          # the Foundation archive is 99% laboratory records
            name = (r.get("description") or "").strip()
            if not name:
                continue
            foods[r["fdc_id"]] = {
                "fdc_id": r["fdc_id"],
                "name": name,
                "source": source_key,
                "category": categories.get(r.get("food_category_id", ""), ""),
                "portions": [],
            }

        # Nutrients. Preference decides ties (kcal beats the Atwater estimate).
        best: dict[str, dict[str, tuple[int, float]]] = {}
        for r in _open_csv(zf, names, "food_nutrient.csv"):
            fid = r["fdc_id"]
            if fid not in foods:
                continue
            hit = lookup.get(r["nutrient_id"])
            if not hit:
                continue
            amount = _num(r.get("amount"))
            if amount is None:
                continue
            key, pref = hit
            slot = best.setdefault(fid, {})
            if key not in slot or pref < slot[key][0]:
                slot[key] = (pref, amount)

        for fid, vals in best.items():
            for key, (_pref, amount) in vals.items():
                foods[fid][key] = round(amount, 3)

        # Branded foods carry their own metadata table, and it's the good stuff:
        # brand, barcode, a real serving size, and a far better category than the
        # 28 generic ones. They also have no food_portion rows at all — the
        # serving lives here instead, so it has to be synthesised.
        for r in _open_csv(zf, names, "branded_food.csv"):
            f = foods.get(r["fdc_id"])
            if not f:
                continue
            brand = (r.get("brand_owner") or r.get("brand_name") or "").strip()
            if brand:
                f["brand"] = brand[:60]
            if r.get("gtin_upc"):
                f["barcode"] = r["gtin_upc"].strip()
            if r.get("branded_food_category"):
                f["category"] = r["branded_food_category"].strip()[:60]
            size = _num(r.get("serving_size"))
            unit = (r.get("serving_size_unit") or "").strip().lower()
            # ml is only grams for water; close enough for a default portion, and
            # the user can edit the weight when adopting.
            if size and 0 < size <= 2000 and unit in ("g", "ml"):
                label = (r.get("household_serving_fulltext") or "").strip()[:40]
                f["portions"].append({
                    "label": label or f"{_fmt(size)} {unit}",
                    "grams": round(size, 2), "seq": 0,
                })

        # Portions — what makes "1 cup" possible instead of "how many grams?"
        for r in _open_csv(zf, names, "food_portion.csv"):
            fid = r["fdc_id"]
            if fid not in foods:
                continue
            grams = _num(r.get("gram_weight"))
            if not grams or grams <= 0:
                continue
            foods[fid]["portions"].append({
                "label": _portion_label(r.get("amount"), units.get(r.get("measure_unit_id", ""), ""),
                                        r.get("portion_description"), r.get("modifier")),
                "grams": round(grams, 2),
                "seq": _num(r.get("seq_num"), 999) or 999,
            })

    out = []
    for f in foods.values():
        # A food with no calories is a laboratory stub, not something you eat.
        if f.get("calories") is None:
            continue
        f["portions"].sort(key=lambda p: p["seq"])
        for p in f["portions"]:
            p.pop("seq", None)
        # Neo4j cannot hold a list of maps as a property.
        f["portions_json"] = json.dumps(f.pop("portions")[:12])
        f["search_name"] = f["name"].lower()
        out.append(f)
    return out


# --------------------------------------------------------------------------- #
# Open Food Facts — the supermarket shelf
#
# A completely different animal from USDA: crowd-sourced, ~4 million products,
# tab-separated, ~9 GB uncompressed, and of very mixed quality. It is the only
# free source with real barcodes and British supermarket products, which is
# exactly what USDA lacks.
#
# Two things it does differently, both of which silently corrupt data if missed:
#   * **Minerals and vitamins are stated in GRAMS**, not mg/µg. Import
#     `sodium_100g` as-is and a bag of crisps reports 0.5 mg of sodium.
#   * **Anyone can edit it.** Values like 3,000,000 kcal/100 g are in there. Every
#     row goes through a sanity check before it is allowed in.
# --------------------------------------------------------------------------- #
OFF_FILENAME = "en.openfoodfacts.org.products.csv.gz"

# canonical key -> (OFF column, multiplier to reach USDA's unit)
OFF_FIELDS: dict[str, tuple[str, float]] = {
    "calories":      ("energy-kcal_100g", 1),
    "protein":       ("proteins_100g", 1),
    "carbs":         ("carbohydrates_100g", 1),
    "fat":           ("fat_100g", 1),
    "fibre":         ("fiber_100g", 1),
    "sugar":         ("sugars_100g", 1),
    "saturated_fat": ("saturated-fat_100g", 1),
    # g -> mg
    "cholesterol":   ("cholesterol_100g", 1000),
    "sodium":        ("sodium_100g", 1000),
    "potassium":     ("potassium_100g", 1000),
    "calcium":       ("calcium_100g", 1000),
    "iron":          ("iron_100g", 1000),
    "magnesium":     ("magnesium_100g", 1000),
    "zinc":          ("zinc_100g", 1000),
    "vitamin_c":     ("vitamin-c_100g", 1000),
    # g -> µg
    "vitamin_a":     ("vitamin-a_100g", 1_000_000),
    "vitamin_d":     ("vitamin-d_100g", 1_000_000),
    "vitamin_b12":   ("vitamin-b12_100g", 1_000_000),
    "folate":        ("folates_100g", 1_000_000),
}

# Nothing edible exceeds ~900 kcal/100 g (pure fat is 900), and no macro can
# exceed 100 g per 100 g. These bounds throw out the typos, not real food.
OFF_LIMITS = {"calories": 900.0, "protein": 100.0, "carbs": 100.0, "fat": 100.0,
              "fibre": 100.0, "sugar": 100.0, "saturated_fat": 100.0}


def off_archive() -> Path | None:
    p = source_dir() / "openfoodfacts" / OFF_FILENAME
    return p if p.is_file() else None


def off_countries() -> list[str]:
    """Which countries to keep, as OFF country tags. Empty list = the whole world."""
    from config import settings
    raw = (getattr(settings, "off_countries", "") or "").strip()
    return [c.strip().lower() for c in raw.split(",") if c.strip()]


def parse_openfoodfacts(path: Path, chunk_size: int = 20000):
    """Stream the OFF export, yielding validated chunks.

    A generator, not a list: the file holds millions of rows and materialising
    them all would cost gigabytes of RAM for no benefit.
    """
    import gzip

    wanted = off_countries()
    # Some OFF fields (ingredient lists) are enormous; the default limit throws.
    csv.field_size_limit(10_000_000)

    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as fh:
        # QUOTE_NONE: the export is unquoted and does contain stray quote
        # characters inside product names — letting csv interpret them merges
        # rows together and produces nonsense.
        reader = csv.DictReader(fh, delimiter="\t", quoting=csv.QUOTE_NONE)
        chunk: list[dict] = []
        for row in reader:
            rec = _off_record(row, wanted)
            if rec is None:
                continue
            chunk.append(rec)
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        if chunk:
            yield chunk


def _off_record(row: dict, wanted: list[str]) -> dict | None:
    code = (row.get("code") or "").strip()
    name = (row.get("product_name") or "").strip()
    if not code or not name or len(name) > 150:
        return None

    if wanted:
        tags = (row.get("countries_tags") or "").lower()
        if not any(c in tags for c in wanted):
            return None

    out: dict = {}
    for key, (col, mult) in OFF_FIELDS.items():
        v = _num(row.get(col))
        if v is None or v < 0:
            continue
        limit = OFF_LIMITS.get(key)
        if limit is not None and v > limit:
            return None                      # a bad row, not a bad field — drop it
        out[key] = round(v * mult, 3)

    # Salt is far more commonly filled in than sodium on European labels.
    if "sodium" not in out:
        salt = _num(row.get("salt_100g"))
        if salt is not None and 0 <= salt <= 100:
            out["sodium"] = round(salt / 2.5 * 1000, 3)

    # Require the four macros. A product with a name and nothing else is not
    # something you can log, and 4 million of those would wreck search.
    if any(out.get(k) is None for k in MACRO_KEYS):
        return None

    brand = (row.get("brands") or "").split(",")[0].strip()[:60]
    category = (row.get("categories_en") or "").split(",")[0].strip()[:60]

    portions = []
    grams = _num(row.get("serving_quantity"))
    if grams and 0 < grams <= 2000:
        label = (row.get("serving_size") or "").strip()[:40] or "1 serving"
        portions.append({"label": label, "grams": round(grams, 2)})

    out.update({
        "fdc_id": "off:" + code,     # namespaced: OFF barcodes would collide with FDC ids
        "name": name,
        "source": "off",
        "category": category,
        "brand": brand,
        "barcode": code,
        "portions_json": json.dumps(portions),
        "search_name": name.lower(),
    })
    return out


# --------------------------------------------------------------------------- #
# CoFID — McCance & Widdowson's The Composition of Foods Integrated Dataset
#
# The UK government's own food composition tables (Public Health England /
# OHID, Open Government Licence). Only ~2,900 foods, but they are *British*
# foods measured in British laboratories — "Digestive biscuits", "Baked beans in
# tomato sauce", "Yorkshire pudding" — which no USDA table will ever contain.
# For a UK user this is the highest-quality whole-food reference available.
#
# It ships as a spreadsheet with the nutrients split across sheets and three
# header rows, so it needs joining on food code. Two conventions to respect:
#   'Tr' means trace — a real measurement of ~0, not missing.
#   'N'  means not measured — genuinely unknown, and must stay absent rather
#        than become a zero, which would read as "contains none of it".
# --------------------------------------------------------------------------- #
COFID_LABEL = "UK CoFID (McCance & Widdowson)"

# sheet -> {canonical key: column header}
COFID_SHEETS: dict[str, dict[str, str]] = {
    "1.3 Proximates": {
        "protein": "Protein (g)", "fat": "Fat (g)", "carbs": "Carbohydrate (g)",
        "calories": "Energy (kcal) (kcal)", "sugar": "Total sugars (g)",
        "fibre": "AOAC fibre (g)", "saturated_fat": "Satd FA /100g fd (g)",
        "cholesterol": "Cholesterol (mg)",
    },
    "1.4 Inorganics": {
        "sodium": "Sodium (mg)", "potassium": "Potassium (mg)", "calcium": "Calcium (mg)",
        "magnesium": "Magnesium (mg)", "iron": "Iron (mg)", "zinc": "Zinc (mg)",
    },
    "1.5 Vitamins": {
        "vitamin_a": "Retinol Equivalent (µg)", "vitamin_d": "Vitamin D (µg)",
        "vitamin_b12": "Vitamin B12 (µg)", "folate": "Folate (µg)",
        "vitamin_c": "Vitamin C (mg)",
    },
}
COFID_DATA_ROW = 4          # rows 1-3 are three stacked header rows


def cofid_archive() -> Path | None:
    d = source_dir() / "cofid"
    if not d.is_dir():
        return None
    files = sorted(d.glob("*.xlsx"))
    return files[-1] if files else None


def _cofid_value(raw) -> float | None:
    """'Tr' is a trace (≈0 and known); 'N' is not measured (unknown, stays out)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s.upper() == "N":
        return None
    if s.lower().startswith("tr"):
        return 0.0
    s = s.lstrip("<~").strip()       # a few values read "<0.1"
    try:
        return float(s)
    except ValueError:
        return None


def parse_cofid(path: Path) -> list[dict]:
    """Read the CoFID workbook into catalogue records, joined on food code."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    foods: dict[str, dict] = {}

    for sheet_name, mapping in COFID_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            log.warning("CoFID: sheet %r missing, skipped", sheet_name)
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        # The Inorganics sheet leaves A1 blank, so the code column is found by
        # position rather than by name.
        idx = {key: header.index(col) for key, col in mapping.items() if col in header}
        for key, col in mapping.items():
            if col not in header:
                log.warning("CoFID: column %r missing from %s", col, sheet_name)

        for i, row in enumerate(rows, start=2):
            if i < COFID_DATA_ROW or not row:
                continue
            code = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not code or not name:
                continue
            f = foods.setdefault(code, {
                "fdc_id": "cofid:" + code,
                "name": name,
                "source": "cofid",
                # CoFID's "Group" is a two-letter internal code (DR, FA, JA)
                # that means nothing outside the manual, so it is deliberately
                # not carried through as a category — the food's name is a far
                # better signal, and that's what _app_category() reads.
                "category": "",
                "portions": [],
            })
            for key, col_i in idx.items():
                if col_i < len(row):
                    v = _cofid_value(row[col_i])
                    if v is not None:
                        f[key] = round(v, 3)

    out = []
    for f in foods.values():
        if any(f.get(k) is None for k in MACRO_KEYS):
            continue                          # not loggable without the four macros
        f["portions_json"] = json.dumps(f.pop("portions"))
        f["search_name"] = f["name"].lower()
        out.append(f)
    return out


def find_archives() -> list[tuple[str, str, str, Path]]:
    """Locate each source archive on disk, newest release wins."""
    src = source_dir()
    found = []
    if not src.is_dir():
        return found
    for key, label, data_type, prefix in SOURCES:
        matches = sorted(p for p in src.glob(f"{prefix}*.zip") if p.is_file())
        if matches:
            found.append((key, label, data_type, matches[-1]))
    return found


# --------------------------------------------------------------------------- #
# Import into Neo4j
# --------------------------------------------------------------------------- #
# No `(:CatalogFood)-[:OWNED_BY]->(:Module)` edge, unlike every other node in the
# app. This is Mainframe-level reference data, not content a module owns, and at
# this scale the edge would cost a million relationships to answer a question
# nobody asks. Dropping it is also most of the reason the import is fast.
WRITE = """
UNWIND $rows AS row
MERGE (c:CatalogFood {fdc_id: row.fdc_id})
SET c += row
"""

BATCH = 5000


async def _write_batches(rows: list[dict], key: str, progress=None) -> None:
    """Write in batches over ONE session.

    `run_write` opens a session per call, which is right for a normal request and
    completely wrong here — session setup dominated the import and held it to
    ~200 rows/s. One session for the whole source is worth ~20×.
    """
    from db import get_driver
    from config import settings

    driver = get_driver()
    async with driver.session(database=settings.neo4j_database) as session:
        for i in range(0, len(rows), BATCH):
            await session.run(WRITE, rows=rows[i:i + BATCH])
            if progress:
                progress(key, min(i + BATCH, len(rows)), len(rows))


async def import_all(progress=None, only: str | None = None) -> dict:
    """(Re)import every source found on disk. Idempotent — MERGEs on fdc_id.

    `only` limits the run to a single source key, so re-importing Open Food Facts
    doesn't mean re-reading half a gigabyte of USDA archives as well.
    """
    archives = find_archives()
    off = off_archive()
    if not archives and not off and not cofid_archive():
        raise FileNotFoundError(
            f"No food archives in {source_dir()}. "
            "USDA: https://fdc.nal.usda.gov/download-datasets.html · "
            "Open Food Facts: https://world.openfoodfacts.org/data"
        )

    summary: dict = {"sources": [], "total": 0}

    for key, label, data_type, path in archives:
        if only and key != only:
            continue
        rows = parse_archive(path, data_type, key)
        await _write_batches(rows, key, progress)
        with_macros = sum(1 for r in rows if all(r.get(k) is not None for k in MACRO_KEYS))
        summary["sources"].append({
            "key": key, "label": label, "archive": path.name,
            "foods": len(rows), "with_full_macros": with_macros,
        })
        summary["total"] += len(rows)
        log.info("food catalogue: %s → %d foods (%d with full macros)", key, len(rows), with_macros)

    cofid = cofid_archive()
    if cofid and (not only or only == "cofid"):
        rows = parse_cofid(cofid)
        await _write_batches(rows, "cofid", progress)
        summary["sources"].append({
            "key": "cofid", "label": COFID_LABEL, "archive": cofid.name,
            "foods": len(rows), "with_full_macros": len(rows),
        })
        summary["total"] += len(rows)
        log.info("food catalogue: cofid → %d foods", len(rows))

    if off and (not only or only == "off"):
        total = 0
        for chunk in parse_openfoodfacts(off):
            await _write_batches(chunk, "off", progress)
            total += len(chunk)
            if progress:
                progress("off", total, 0)          # 0 = total unknown while streaming
        summary["sources"].append({
            "key": "off", "label": OFF_LABEL, "archive": off.name, "foods": total,
            "with_full_macros": total,             # the parser requires all four
        })
        summary["total"] += total
        log.info("food catalogue: open food facts → %d foods", total)

    await _refresh_counts()      # the cached breakdown is now stale
    return summary


async def count() -> int:
    """Total rows. A bare label count uses Neo4j's count store, so this is
    instant even at millions — unlike counting *grouped by source*."""
    from db import run_read
    rows = await run_read("MATCH (c:CatalogFood) RETURN count(c) AS n")
    return rows[0]["n"] if rows else 0


async def _refresh_counts() -> list[dict]:
    """Recount per source and cache it on a singleton node.

    `count(c)` grouped by `c.source` is a full scan of every catalogue node —
    several seconds once there are millions, and the Food Database tab asks for
    it on every render. Counting once at import time and reading a cached number
    afterwards is the difference between a tab that opens instantly and one that
    hangs.
    """
    from db import run_read, run_write
    rows = await run_read(
        "MATCH (c:CatalogFood) RETURN c.source AS source, count(c) AS n ORDER BY n DESC"
    )
    counts = [{"source": r["source"], "n": r["n"]} for r in rows]
    await run_write(
        "MERGE (m:CatalogMeta {id:'catalog'}) SET m.counts = $json, m.counted_at = $at",
        json=json.dumps(counts), at=_now_iso(),
    )
    return counts


def _now_iso() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


async def status() -> dict:
    """What's in the catalogue right now, and what's on disk to import."""
    from db import run_read
    cached = await run_read("MATCH (m:CatalogMeta {id:'catalog'}) RETURN m{.*} AS m")
    if cached and cached[0]["m"].get("counts"):
        rows = json.loads(cached[0]["m"]["counts"])
        counted_at = cached[0]["m"].get("counted_at")
    else:
        rows = await _refresh_counts()
        counted_at = _now_iso()
    labels = {k: lbl for k, lbl, _dt, _p in SOURCES}
    labels["off"] = OFF_LABEL
    labels["cofid"] = COFID_LABEL
    archives = [{"key": k, "label": lbl, "file": p.name,
                 "size_mb": round(p.stat().st_size / 1e6, 1)}
                for k, lbl, _dt, p in find_archives()]
    for key, label, path in (("off", OFF_LABEL, off_archive()),
                             ("cofid", COFID_LABEL, cofid_archive())):
        if path:
            archives.append({"key": key, "label": label, "file": path.name,
                             "size_mb": round(path.stat().st_size / 1e6, 1)})
    return {
        "total": sum(r["n"] for r in rows),
        "sources": [{"key": r["source"], "label": labels.get(r["source"], r["source"]),
                     "foods": r["n"]} for r in rows],
        "archives": archives,
        "source_dir": str(source_dir()),
        "off_countries": off_countries() or ["everywhere"],
        "counted_at": counted_at,
    }


if __name__ == "__main__":       # python foodcatalog.py — one-shot import
    import asyncio

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    async def main():
        def show(key, done, total):
            print(f"  {key}: {done}/{total}", end="\r", flush=True)
        result = await import_all(progress=show)
        print("\n" + json.dumps(result, indent=2))

    asyncio.run(main())
